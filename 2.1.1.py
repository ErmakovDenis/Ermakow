import csv
import math
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side

name_list = ["name", "salary_from", "salary_to", "salary_currency", "area_name", "published_at"]

currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
                   "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}


class Vacancy:
    """Класс для получения данных о вакансии.

    Attributes:
        name (str): Название вакансии
        salary_average (int): Средняя зарплата в рублях
        area_name (str): Название города
        publication_year (int): Год публикации вакансии
    """

    def __init__(self, vacancies):
        """Инициализирует объект Vacancy, вычисляет среднюю зарплату и переводит в рубли

        Args:
            vacancies (dict): Вакансия
        """
        self.name = vacancies[name_list[0]]
        self.salary_average = math.floor((float(vacancies[name_list[1]]) + float(vacancies[name_list[2]])) / 2) \
                              * currency_to_rub[vacancies[name_list[3]]]
        self.area_name = vacancies[name_list[4]]
        self.publication_year = int(vacancies[name_list[5]][:4])


class DataSet:
    """Класс для получения и печати статистик.

    Attributes:
        filename (str): Название файла с данными о вакансиях
        vacancy_name (str): Название выбранной профессии
    """

    def __init__(self, filename, vacancy_name):
        """Инициализирует объект DataSet.

        Args:
            filename (str): Название файла с данными о вакансиях
            vacancy_name (str): Название выбранной профессии
        """
        self.filename, self.vacancy_name = filename, vacancy_name

    def csv_reader(self):
        """Считывает данные из входного файла, получает все необходимые статистики для дальнейшей работы и печатает их

        Returns:
            dict, dict, dict, dict, dict, dict: Все необходимые статистики
        """
        with open(self.filename, mode='r', encoding='utf-8-sig') as file:
            count = 0
            salary = {}
            city = {}
            number = {}
            vacancy_number = {}
            salary_of_name = {}
            number_of_name = {}
            header = []
            reader = csv.reader(file)
            for index, row in enumerate(reader):
                if index == 0:
                    csv_header_length = len(row)
                    header = row
                elif '' not in row and len(row) == csv_header_length:
                    vacancies = Vacancy(dict(zip(header, row)))
                    if vacancies.publication_year not in salary:
                        salary[vacancies.publication_year] = [vacancies.salary_average]
                    else:
                        salary[vacancies.publication_year].append(vacancies.salary_average)

                    if vacancies.area_name not in city:
                        city[vacancies.area_name] = [vacancies.salary_average]
                    else:
                        city[vacancies.area_name].append(vacancies.salary_average)

                    if vacancies.area_name in number:
                        number[vacancies.area_name] += 1
                    else:
                        number[vacancies.area_name] = 1

                    if vacancies.publication_year in vacancy_number:
                        vacancy_number[vacancies.publication_year] += 1
                    else:
                        vacancy_number[vacancies.publication_year] = 1

                    if vacancies.name.find(self.vacancy_name) != -1:
                        if vacancies.publication_year not in salary_of_name:
                            salary_of_name[vacancies.publication_year] = [vacancies.salary_average]
                        else:
                            salary_of_name[vacancies.publication_year].append(vacancies.salary_average)

                        if vacancies.publication_year in number_of_name:
                            number_of_name[vacancies.publication_year] += 1
                        else:
                            number_of_name[vacancies.publication_year] = 1
                    count += 1

        if not salary_of_name:
            number_of_name = dict([(k, 0) for k, v in vacancy_number.copy().items()])
            salary_of_name = dict([(k, []) for k, v in salary.copy().items()])

        dynamics1, dynamics2, dynamics3, dynamics4 = {}, {}, {}, {}
        for year, salaries in salary.items():
            dynamics1[year] = int(sum(salaries) / len(salaries))

        for year, salaries in salary_of_name.items():
            dynamics2[year] = 0 if len(salaries) == 0 else int(sum(salaries) / len(salaries))

        for year, salaries in city.items():
            dynamics3[year] = int(sum(salaries) / len(salaries))

        for year, salaries in number.items():
            dynamics4[year] = round(salaries / count, 4)

        dynamics4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in dynamics4.items()]))
        dynamics4.sort(key=lambda a: a[-1], reverse=True)
        dynamics3 = list(
            filter(lambda a: a[0] in list(dict(dynamics4).keys()), [(key, value) for key, value in dynamics3.items()]))
        dynamics3.sort(key=lambda a: a[-1], reverse=True)

        print('Динамика уровня зарплат по годам: {0}'.format(dynamics1))
        print('Динамика количества вакансий по годам: {0}'.format(vacancy_number))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(dynamics2))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(number_of_name))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(dict(dynamics3[:10])))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(dict(dynamics4.copy()[:10])))

        return dynamics1, vacancy_number, dynamics2, number_of_name, dict(dynamics3[:10]), dict(dynamics4.copy()[:10])


class InputConnect:
    """Класс для получения объектов DataSet и Report.

    Attributes:
        filename (str): Название файла с данными о вакансиях
        name_vacancy (str): Название выбранной профессии
    """

    def __init__(self):
        """Инициализирует объект InputConnect.
        """
        self.filename, self.name_vacancy = input('Введите название файла: '), input('Введите название профессии: ')

        dataset = DataSet(self.filename, self.name_vacancy)
        dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6 = dataset.csv_reader()

        report = Report(self.name_vacancy, dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6)
        report.generate_excel()


class Report:
    """Класс для получения отчета по полученным динамикам.

    Attributes:
        workbook (Workbook()): Рабочая книга для получания xlsx-файла
        name_vacancy (str): Название выбранной профессии
        dynamics1 (dict): Динамика уровня зарплат по годам
        dynamics2 (dict): Динамика количества вакансий по годам
        dynamics3 (dict): Динамика уровня зарплат по годам для выбранной профессии
        dynamics4 (dict): Динамика количества вакансий по годам для выбранной профессии
        dynamics5 (dict): Уровень зарплат по городам (в порядке убывания)
        dynamics6 (dict): Доля вакансий по городам (в порядке убывания)
    """

    def __init__(self, name_vacancy, dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6):
        """Инициализирует объект Report и получает объект Workbook.

        Args:
            name_vacancy (str): Название выбранной профессии
            dynamics1 (dict): Динамика уровня зарплат по годам
            dynamics2 (dict): Динамика количества вакансий по годам
            dynamics3 (dict): Динамика уровня зарплат по годам для выбранной профессии
            dynamics4 (dict): Динамика количества вакансий по годам для выбранной профессии
            dynamics5 (dict): Уровень зарплат по городам (в порядке убывания)
            dynamics6 (dict): Доля вакансий по городам (в порядке убывания)
        """
        self.workbook = Workbook()
        self.name_vacancy = name_vacancy
        self.dynamics1, self.dynamics2, self.dynamics3, self.dynamics4, self.dynamics5, self.dynamics6 \
            = dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6

    def generate_excel(self):
        """Генерирует две страницы в рабочей книге с статистикой по годам и статистикой по городам, после чего сохраняет рабочую книгу в файл report.xlsx
        """
        work_sheet1 = self.workbook.active
        work_sheet1.title = 'Статистика по годам'
        work_sheet1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.name_vacancy, 'Количество вакансий',
                            'Количество вакансий - ' + self.name_vacancy])
        for year in self.dynamics1.keys():
            work_sheet1.append(
                [year, self.dynamics1[year], self.dynamics3[year], self.dynamics2[year], self.dynamics4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.name_vacancy, ' Количество вакансий',
                 ' Количество вакансий - ' + self.name_vacancy]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            work_sheet1.column_dimensions[get_column_letter(i)].width = column_width + 2

        data = []
        data.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (city1, value1), (city2, value2) in zip(self.dynamics5.items(), self.dynamics6.items()):
            data.append([city1, value1, '', city2, value2])
        work_sheet2 = self.workbook.create_sheet('Статистика по городам')
        for row in data:
            work_sheet2.append(row)

        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            work_sheet2.column_dimensions[get_column_letter(i)].width = column_width + 2

        bold = Font(bold=True)
        for column in 'ABCDE':
            work_sheet1[column + '1'].font = bold
            work_sheet2[column + '1'].font = bold

        for index, _ in enumerate(self.dynamics5):
            work_sheet2['E' + str(index + 2)].number_format = '0.00%'

        slim = Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for column in 'ABDE':
                work_sheet2[column + str(row + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)

        self.dynamics1[1] = 1
        for row, _ in enumerate(self.dynamics1):
            for column in 'ABCDE':
                work_sheet1[column + str(row + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)

        self.workbook.save('report.xlsx')


if __name__ == '__main__':
    InputConnect()