import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pathlib
import pdfkit
import math

currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                   "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}
name_list = ["name", "salary_from", "salary_to", "salary_currency", "area_name", "published_at"]

list_print1 = ['Динамика уровня зарплат по годам: ', 'Динамика количества вакансий по годам: ',
               'Динамика уровня зарплат по годам для выбранной профессии: ',
               'Динамика количества вакансий по годам для выбранной профессии: ',
               'Уровень зарплат по городам (в порядке убывания): ', 'Доля вакансий по городам (в порядке убывания): ']


class Vacancy:
    """Класс для получения данных о вакансии.

    Attributes:
        name (str): Название вакансии
        salary_average (int): Средняя зарплата в рублях
        area_name (str): Название города
        publication_year (int): Год публикации вакансии
    """

    def __init__(self, vacancies):
        """Инициализирует объект Vacancy, вычисляет среднюю зарплату и переводит в рубли

        Args:
            vacancies (dict): Вакансия
        """
        self.name = vacancies[name_list[0]]
        self.salary_average = math.floor((float(vacancies[name_list[1]]) + float(vacancies[name_list[2]])) / 2) \
                              * currency_to_rub[vacancies[name_list[3]]]
        self.area_name = vacancies[name_list[4]]
        self.publication_year = int(vacancies[name_list[5]][:4])


class DataSet:
    """Класс для получения и печати статистик.

    Attributes:
        filename (str): Название файла с данными о вакансиях
        vacancy_name (str): Название выбранной профессии
    """

    def __init__(self, filename, vacancy_name):
        """Инициализирует объект DataSet.

        Args:
            filename (str): Название файла с данными о вакансиях
            vacancy_name (str): Название выбранной профессии
        """
        self.filename, self.vacancy_name = filename, vacancy_name

    def csv_reader(self):
        """Считывает данные из входного файла

        Returns:
            dict: Все вакансии с информацией о них.
        """
        with open(self.filename, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            header_length = len(header)
            for row in reader:
                if '' not in row and len(row) == header_length:
                    yield dict(zip(header, row))

    @staticmethod
    def average(dict):
        """Высчитывает среднее значение.

        Args:
            dict (dict): Словарь с значениями
        Returns:
            dict: Словарь с обновленными, средними значениями
        """
        new_dict = {}
        for k, v in dict.items():
            new_dict[k] = int(sum(v) / len(v))
        return new_dict

    @staticmethod
    def increment(dict, k, с):
        """Дополняет словарь всех средних зарплат за год или в городе

        Args:
            dict (dict): Словарь со всеми зарплатами за год или в городе по вакансии
            k (int): Год или город вакансии, по которому идет подсчет
            с (list): Средняя зарплата у вакансии
        """
        if k in dict:
            dict[k] += с
        else:
            dict[k] = с

    def get_dynamics(self):
        """Получает все необходимые статистики для дальнейшей работы

        Returns:
            dict, dict, dict, dict, dict, dict: Все необходимые статистики
        """
        salary = {}
        salary_of_name = {}
        city = {}
        count = 0

        for vacancy_dictionary in self.csv_reader():
            vacancy = Vacancy(vacancy_dictionary)
            self.increment(salary, vacancy.publication_year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.increment(salary_of_name, vacancy.publication_year, [vacancy.salary_average])
            self.increment(city, vacancy.area_name, [vacancy.salary_average])
            count += 1

        number = dict([(k, len(v)) for k, v in salary_of_name.items()])
        vacancy_number = dict([(k, len(v)) for k, v in salary.items()])

        if not salary_of_name:
            number = dict([(k, 0) for k, v in vacancy_number.items()])
            salary_of_name = dict([(k, [0]) for k, v in salary.items()])

        dynamics1, dynamics2, dynamics3 = self.average(salary), self.average(salary_of_name), self.average(city)

        dynamics4 = {}
        for y, s in city.items():
            dynamics4[y] = round(len(s) / count, 4)
        dynamics4 = list(filter(lambda x: x[-1] >= 0.01, [(k, v) for k, v in dynamics4.items()]))
        dynamics4.sort(key=lambda x: x[-1], reverse=True)
        dynamics5 = dict(dynamics4.copy()[:10])
        dynamics4 = dict(dynamics4)
        dynamics3 = list(filter(lambda x: x[0] in list(dynamics4.keys()), [(k, v) for k, v in dynamics3.items()]))
        dynamics3.sort(key=lambda x: x[-1], reverse=True)
        dynamics3 = dict(dynamics3[:10])

        return dynamics1, vacancy_number, dynamics2, number, dynamics3, dynamics5

    @staticmethod
    def print_statistic(dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6):
        """Выводит все динамики с описанием

        Args:
            dynamics1 (dict): Динамика уровня зарплат по годам
            dynamics2 (dict): Динамика количества вакансий по годам
            dynamics3 (dict): Динамика уровня зарплат по годам для выбранной профессии
            dynamics4 (dict): Динамика количества вакансий по годам для выбранной профессии
            dynamics5 (dict): Уровень зарплат по городам (в порядке убывания)
            dynamics6 (dict): Доля вакансий по городам (в порядке убывания)
        Prints:
            Печатать каждой динамики с описанием
        """
        list_print2 = [dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6]
        for i in range(len(list_print1)):
            print(list_print1[i] + '{0}'.format(list_print2[i]))


class InputConnect:
    """Класс для получения объектов DataSet и Report.

    Attributes:
        filename (str): Название файла с данными о вакансиях
        name_vacancy (str): Название выбранной профессии
    """

    def __init__(self):
        """Инициализирует объект InputConnect.
        """
        self.filename = input('Введите название файла: ')
        self.name_vacancy = input('Введите название профессии: ')

        dataset = DataSet(self.filename, self.name_vacancy)

        dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6 = dataset.get_dynamics()
        dataset.print_statistic(dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6)
        new_graphic = Report(self.name_vacancy, dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6)
        new_graphic.generate_image()
        work_sheet1 = new_graphic.get_first_sheet()
        work_sheet2, len_new_data = new_graphic.get_second_sheet()
        new_graphic.generate_excel(work_sheet1, work_sheet2, len_new_data)
        new_graphic.generate_pdf()


class Report:
    """Класс для получения отчета по полученным динамикам.

    Attributes:
        workbook (Workbook()): Рабочая книга для получания xlsx-файла
        name_vacancy (str): Название выбранной профессии
        dynamics1 (dict): Динамика уровня зарплат по годам
        dynamics2 (dict): Динамика количества вакансий по годам
        dynamics3 (dict): Динамика уровня зарплат по годам для выбранной профессии
        dynamics4 (dict): Динамика количества вакансий по годам для выбранной профессии
        dynamics5 (dict): Уровень зарплат по городам (в порядке убывания)
        dynamics6 (dict): Доля вакансий по городам (в порядке убывания)
    """

    def __init__(self, name_vacancy, dynamics1, dynamics2, dynamics3, dynamics4, dynamics5, dynamics6):
        """Инициализирует объект Report и получает объект Workbook.

        Args:
            name_vacancy (str): Название выбранной профессии
            dynamics1 (dict): Динамика уровня зарплат по годам
            dynamics2 (dict): Динамика количества вакансий по годам
            dynamics3 (dict): Динамика уровня зарплат по годам для выбранной профессии
            dynamics4 (dict): Динамика количества вакансий по годам для выбранной профессии
            dynamics5 (dict): Уровень зарплат по городам (в порядке убывания)
            dynamics6 (dict): Доля вакансий по городам (в порядке убывания)
        """
        self.workbook = Workbook()
        self.name_vacancy = name_vacancy
        self.dynamics1 = dynamics1
        self.dynamics2 = dynamics2
        self.dynamics3 = dynamics3
        self.dynamics4 = dynamics4
        self.dynamics5 = dynamics5
        self.dynamics6 = dynamics6

    def generate_image(self):
        """Генерирует 4 диаграммы в на одной старнице на основе статистик, после чего сохраняет картинку в файл graph.png
        """
        x = np.arange(len(self.dynamics1.keys()))
        width = 0.35

        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(ncols=2, nrows=2)
        ax1.bar(x - width / 2, self.dynamics1.values(), width, label='средняя з/п')
        ax1.bar(x + width / 2, self.dynamics3.values(), width, label='з/п {0}'.format(self.name_vacancy))
        plt.rcParams['font.size'] = '8'
        ax1.set_title('Уровень зарплат по годам')
        ax1.set_xticks(x, self.dynamics1.keys(), rotation=90)
        ax1.legend(fontsize=8)
        ax1.grid(axis='y')

        ax2.bar(x - width / 2, self.dynamics2.values(), width, label='количество вакансий')
        ax2.bar(x + width / 2, self.dynamics4.values(), width,
                label='количество вакансий {0}'.format(self.name_vacancy))
        ax2.set_xticks(x, self.dynamics2.keys(), rotation=90)
        ax2.set_title('Количество вакансий по годам')
        ax2.legend(fontsize=8)
        ax2.grid(axis='y')
        fig.tight_layout()

        areas = []
        for area in self.dynamics5.keys():
            areas.append(str(area).replace(' ', '\n').replace('-', '-\n'))
        y_pos = np.arange(len(areas))
        performance = self.dynamics5.values()
        error = np.random.rand(len(areas))
        ax3.barh(y_pos, performance, xerr=error, align='center')
        ax3.set_title('Уровень зарплат по мегаполисам')
        ax3.set_yticks(y_pos, labels=areas, size=6)
        ax3.invert_yaxis()
        ax3.grid(axis='x')

        val = list(self.dynamics6.values()) + [1 - sum(list(self.dynamics6.values()))]
        k = list(self.dynamics6.keys()) + ['Другие']
        ax4.pie(val, labels=k, startangle=170)
        ax4.set_title('Доля вакансий по городам')

        plt.tight_layout()
        plt.savefig('graph.png', dpi=300)

    def get_first_sheet(self):
        """Генерирует первую страницу рабочей книги с статистикой по годам.

        Returns:
            Worksheet: Первая страница рабочей книги
        """
        work_sheet1 = self.workbook.active
        work_sheet1.title = 'Статистика по годам'
        work_sheet1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.name_vacancy, 'Количество вакансий',
                            'Количество вакансий - ' + self.name_vacancy])
        for year in self.dynamics1.keys():
            work_sheet1.append(
                [year, self.dynamics1[year], self.dynamics3[year], self.dynamics2[year], self.dynamics4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.name_vacancy, ' Количество вакансий',
                 ' Количество вакансий - ' + self.name_vacancy]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            work_sheet1.column_dimensions[get_column_letter(i)].width = column_width + 2
        return work_sheet1

    def get_second_sheet(self):
        """Генерирует вторую страницу рабочей книги с статистикой по городам.

        Returns:
            Worksheet: Вторая страница рабочей книги
            int: Количество строк
        """
        new_data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]

        for (city1, value1), (city2, value2) in zip(self.dynamics5.items(), self.dynamics6.items()):
            new_data.append([city1, value1, '', city2, value2])
        work_sheet2 = self.workbook.create_sheet('Статистика по городам')
        for r in new_data:
            work_sheet2.append(r)

        column_widths = []
        for r in new_data:
            for i, cell in enumerate(r):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):
            work_sheet2.column_dimensions[get_column_letter(i)].width = column_width + 2

        return work_sheet2, len(new_data)

    def generate_excel(self, work_sheet1, work_sheet2, len_new_data):
        """Сохраняет рабочую книгу в файл report.xlsx
        Args:
            work_sheet1 (Worksheet): Первая страница рабочей книги
            work_sheet2 (Worksheet): Вторая страница рабочей книги
            len_new_data (int): Количество строк
        """
        bold = Font(bold=True)
        for c in 'ABCDE':
            work_sheet1[c + '1'].font = bold
            work_sheet2[c + '1'].font = bold

        for index, _ in enumerate(self.dynamics5):
            work_sheet2['E' + str(index + 2)].number_format = '0.00%'

        slim = Side(border_style='thin', color='00000000')

        for row in range(len_new_data):
            for column in 'ABDE':
                work_sheet2[column + str(row + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)

        self.dynamics1[1] = 1
        for row, _ in enumerate(self.dynamics1):
            for column in 'ABCDE':
                work_sheet1[column + str(row + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)

        self.workbook.save('report.xlsx')

    def generate_pdf(self):
        """Генирирует и сохраняет файл report.pdf, в котором хранятся report.xlsx и graph.png
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        dynamics = []
        for year in self.dynamics2.keys():
            dynamics.append(
                [year, self.dynamics1[year], self.dynamics2[year], self.dynamics3[year], self.dynamics4[year]])

        for key in self.dynamics6:
            self.dynamics6[key] = round(self.dynamics6[key] * 100, 2)

        pdf_template = template.render({'name': self.name_vacancy,
                                        'path': '{0}/{1}'.format(pathlib.Path(__file__).parent.resolve(), 'graph.png'),
                                        'dynamics': dynamics, 'dynamics5': self.dynamics5, 'dynamics6': self.dynamics6})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


if __name__ == '__main__':
    InputConnect()